from django.shortcuts import render
from django.views.generic import ListView
from .models import Carousel,Event,Ministries,Head,Blog,Media

from openpyxl import load_workbook
from .models import Data

def import_data(request):
    wb = load_workbook('data.xls')
    ws = wb.active
    for row in ws.iter_rows(values_only=True):
        for cell in row:
            data = Data(value=cell)
            data.save()
    return HttpResponse("Data imported from Excel sheet to database.")

def home(request) :
    carousel = Carousel.objects.all()
    events = Event.objects.all()
    
    context = {
         "events" : events,
         "carousel" : carousel,
    }

    return render(request, 'sunshine/home.html',context)


def index(request) :
    ministries = Ministries.objects.all()[:6]
    head = Head.objects.all()
    events = Event.objects.all()[:2]

    context = {
         "ministries" : ministries,
         "head" : head,
         "events" : events
    }

    return render(request, 'sunshine/index.html',context)

def blogs(request):
     blog = Blog.objects.all()#(blog,lookup=slug)
     
     context ={
          "blog" : blog
     }

     return render(request, 'sunshine/blogs.html',context)

def post(request):
     blog = Blog.objects.all()[:1]
     
     context ={
          "blog" : blog
     }

     return render(request, 'sunshine/post.html', context)

def media(request):
     media = Media.objects.all()

     context ={
          "media" : media
     }

     return render(request,'sunshine/media.html')

def fetch(request):
    if request.method == 'GET':
        media = request.GET['media']
        return HttpResponse('success')
    else:
        return HttpResponse("unsuccesful")